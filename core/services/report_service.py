from __future__ import annotations

import io
import os
from datetime import datetime, timezone
from functools import lru_cache
from pathlib import Path
from typing import Any

import pandas as pd
from reportlab.lib.pagesizes import A4
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas

from core.models import SimulationResult


# =============================================================================
# Fonts
# =============================================================================
# 운영/컨테이너 환경에서 폰트 경로를 고정하고 싶으면 아래 env를 쓰면 됩니다.
# 예) OKSSANGIMONG_KOREAN_FONT_PATH=/app/assets/fonts/NanumGothic.ttf
ENV_KOREAN_FONT_PATH = "OKSSANGIMONG_KOREAN_FONT_PATH"
ENV_FONT_DIR = "OKSSANGIMONG_FONT_DIR"  # 추가 폰트 디렉토리 지정용(선택)

# ReportLab 내장 CID 폰트 중 한국어는 2개가 사실상 전부입니다.
KOREAN_CID_FONT_CANDIDATES: tuple[str, ...] = (
    "HYSMyeongJo-Medium",
    "HYGothic-Medium",
)

# 파일시스템에서 찾을만한 한글 폰트 후보 (이름, 파일명 후보들)
# - 여기서 찾으면 TTFont로 등록되어 PDF에 임베딩되므로 가장 안정적입니다.
KOREAN_TTF_CANDIDATES: tuple[tuple[str, tuple[str, ...]], ...] = (
    ("NanumGothic", ("NanumGothic.ttf",)),
    ("NanumMyeongjo", ("NanumMyeongjo.ttf",)),
    ("NanumBarunGothic", ("NanumBarunGothic.ttf",)),
    ("NotoSansKR", ("NotoSansKR-Regular.otf", "NotoSansKR-Regular.ttf")),
    ("NotoSerifKR", ("NotoSerifKR-Regular.otf", "NotoSerifKR-Regular.ttf")),
    # Windows/macOS에서 돌아갈 가능성까지 고려(파일이 실제로 있으면 사용)
    ("MalgunGothic", ("malgun.ttf", "Malgun.ttf")),
    ("AppleSDGothicNeo", ("AppleSDGothicNeo.ttc", "AppleSDGothicNeo.ttf")),
)


def _utc_now_iso() -> str:
    return (
        datetime.now(timezone.utc)
        .replace(microsecond=0)
        .isoformat()
        .replace("+00:00", "Z")
    )


def _candidate_font_dirs() -> list[Path]:
    """프로젝트/시스템에서 폰트가 있을 법한 디렉토리 후보를 모읍니다."""
    dirs: list[Path] = []

    # 1) env로 명시한 폰트 디렉토리(선택)
    env_dir = os.getenv(ENV_FONT_DIR)
    if env_dir:
        dirs.append(Path(env_dir).expanduser())

    # 2) 현재 파일 기준으로 상위 디렉토리들을 훑어서 fonts 폴더 후보 추가
    # (레포 구조가 달라도 어느 정도 찾게끔)
    here = Path(__file__).resolve()
    for parent in list(here.parents)[:6]:
        dirs.extend(
            [
                parent / "assets" / "fonts",
                parent / "static" / "fonts",
                parent / "resources" / "fonts",
                parent / "fonts",
            ]
        )

    # 3) 시스템 폰트 디렉토리(리눅스/맥/윈 일부)
    dirs.extend(
        [
            Path("/usr/share/fonts"),
            Path("/usr/local/share/fonts"),
            Path("/usr/share/fonts/truetype"),
            Path("/usr/share/fonts/opentype"),
            Path("/System/Library/Fonts"),  # macOS
            Path("/Library/Fonts"),  # macOS
            Path("C:/Windows/Fonts"),  # Windows
        ]
    )

    # 존재하는 디렉토리만 unique하게 반환
    uniq: list[Path] = []
    seen: set[str] = set()
    for d in dirs:
        if d.exists() and d.is_dir():
            key = str(d.resolve())
            if key not in seen:
                seen.add(key)
                uniq.append(d)
    return uniq


def _try_register_ttf(font_name: str, font_path: Path) -> str | None:
    """TTF/OTF/TTC 폰트를 등록 시도. 성공하면 등록된 폰트 이름 반환."""
    if not (font_path.exists() and font_path.is_file()):
        return None

    registered = set(pdfmetrics.getRegisteredFontNames())
    if font_name in registered:
        return font_name

    try:
        # TTC(폰트 컬렉션)는 첫 번째 서브폰트로 시도(환경에 따라 다를 수 있음)
        if font_path.suffix.lower() == ".ttc":
            pdfmetrics.registerFont(TTFont(font_name, str(font_path), subfontIndex=0))
        else:
            pdfmetrics.registerFont(TTFont(font_name, str(font_path)))
        return font_name
    except Exception:
        return None


@lru_cache(maxsize=1)
def _ensure_korean_font() -> str:
    """
    Hangul이 깨지지 않도록 폰트를 확보합니다.

    우선순위:
      1) ENV_KOREAN_FONT_PATH (TTF/OTF/TTC) 있으면 임베딩 폰트 사용(가장 안정적)
      2) 파일시스템 탐색으로 한글 폰트 찾아 임베딩(TTFont)
      3) ReportLab 내장 CID 한글 폰트(뷰어/환경 의존 가능)
      4) Helvetica (최후)
    """
    # 1) 명시 경로 우선
    env_path = os.getenv(ENV_KOREAN_FONT_PATH)
    if env_path:
        p = Path(env_path).expanduser()
        picked = _try_register_ttf("OkssangimongKorean", p)
        if picked:
            return picked

    # 2) 디렉토리 탐색으로 임베딩 폰트 찾기
    font_dirs = _candidate_font_dirs()
    for font_name, filenames in KOREAN_TTF_CANDIDATES:
        for d in font_dirs:
            for fn in filenames:
                p = d / fn
                picked = _try_register_ttf(font_name, p)
                if picked:
                    return picked

        # 추가로 하위 디렉토리(예: /usr/share/fonts/truetype/nanum)도 있을 수 있으니
        # 너무 무겁지 않게 2레벨까지만 살짝 탐색
        for d in font_dirs:
            try:
                for sub in d.iterdir():
                    if not sub.is_dir():
                        continue
                    for fn in filenames:
                        p = sub / fn
                        picked = _try_register_ttf(font_name, p)
                        if picked:
                            return picked
            except Exception:
                continue

    # 3) CID 폰트 fallback (한글 CID 폰트 2개만)
    for cid_name in KOREAN_CID_FONT_CANDIDATES:
        if cid_name in pdfmetrics.getRegisteredFontNames():
            return cid_name
        try:
            pdfmetrics.registerFont(UnicodeCIDFont(cid_name))
            return cid_name
        except Exception:
            continue

    # 4) 최후의 fallback
    return "Helvetica"


# =============================================================================
# Domain labels
# =============================================================================
GREENING_LABELS = {
    "grass": "잔디",
    "sedum": "세덤",
    "shrub": "관목",
    "tree": "나무",
}


def _greening_label(code: str | None) -> str:
    if not code:
        return "-"
    return GREENING_LABELS.get(code, code)


def _as_number(value: Any) -> Any:
    """Excel로 내보낼 때 Decimal 등도 안전하게 숫자로 최대한 변환."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    try:
        return float(value)
    except Exception:
        return value


# =============================================================================
# Service
# =============================================================================
class ReportService:
    """Create PDF/Excel artifacts from SimulationResult.

    MVP 버전: 간단한 PDF/Excel 생성.
    추후 템플릿/디자인은 UI팀 스타일에 맞춰 개선 가능.
    """

    def build_pdf(self, result: SimulationResult) -> tuple[bytes, str]:
        buf = io.BytesIO()
        c = canvas.Canvas(buf, pagesize=A4)
        w, h = A4  # noqa: F841  # (w는 향후 레이아웃 확장 때 사용 가능)
        font_name = _ensure_korean_font()
        generated_at = _utc_now_iso()

        # PDF 메타데이터(선택)
        try:
            c.setTitle("옥상이몽 시뮬레이션 리포트 (MVP)")
        except Exception:
            pass

        left = 60
        y = h - 60

        def line(text: str, size: int = 11, dy: int = 16) -> None:
            nonlocal y
            if y < 60:
                c.showPage()
                y = h - 60
                c.setFont(font_name, size)
            c.setFont(font_name, size)
            c.drawString(left, y, text)
            y -= dy

        # Header
        line("옥상이몽 시뮬레이션 리포트 (MVP)", size=16, dy=24)
        line(f"Generated (UTC): {generated_at}", size=10, dy=18)
        line(
            f"Engine: {result.engine_version} | Coeff set: {result.coefficient_set_version}",
            size=10,
            dy=30,
        )

        # Inputs
        line("입력값", size=12, dy=18)
        line(f"- 옥상 면적(㎡): {result.roof_area_m2:,.2f}")
        line(f"- 녹화 유형: {_greening_label(result.greening_type)}")
        line(f"- 녹화 비율: {result.coverage_ratio:.2%}", dy=26)

        # Outputs
        line("결과", size=12, dy=18)
        line(f"- 녹화 면적(㎡): {result.green_area_m2:,.2f}")
        line(f"- CO₂ 흡수(kg/년): {result.co2_absorption_kg_per_year:,.2f}")
        line(f"- 온도 저감(℃): {result.temp_reduction_c:,.2f}")
        line(
            f"- 표면온도(전/후): {result.baseline_surface_temp_c:.1f}℃ → {result.after_surface_temp_c:.1f}℃"
        )
        # int일 가능성이 높아서 천단위 콤마
        try:
            trees = f"{int(result.tree_equivalent_count):,}"
        except Exception:
            trees = str(result.tree_equivalent_count)
        line(f"- 소나무 환산(그루): {trees}", dy=30)

        # Footer note
        line("※ 본 리포트는 MVP 산출물이며, 실제 정책/심사 제출 전 계수·근거 검증이 필요합니다.", size=9, dy=14)

        c.showPage()
        c.save()

        pdf_bytes = buf.getvalue()
        filename = "okssangimong_report.pdf"
        return pdf_bytes, filename

    def build_excel(self, result: SimulationResult) -> tuple[bytes, str]:
        buf = io.BytesIO()
        generated_at = _utc_now_iso()

        inputs = pd.DataFrame(
            [
                {"항목": "옥상 면적(㎡)", "값": _as_number(result.roof_area_m2)},
                {"항목": "녹화 유형", "값": _greening_label(result.greening_type)},
                {"항목": "녹화 비율", "값": _as_number(result.coverage_ratio)},  # 숫자로 유지(퍼센트 서식은 아래에서 적용)
                {"항목": "기준 표면 온도(℃)", "값": _as_number(result.baseline_surface_temp_c)},
            ]
        )

        outputs = pd.DataFrame(
            [
                {"항목": "녹화 면적(㎡)", "값": _as_number(result.green_area_m2)},
                {"항목": "CO₂ 흡수(kg/년)", "값": _as_number(result.co2_absorption_kg_per_year)},
                {"항목": "온도 저감(℃)", "값": _as_number(result.temp_reduction_c)},
                {"항목": "녹화 후 표면 온도(℃)", "값": _as_number(result.after_surface_temp_c)},
                {"항목": "소나무 환산(그루)", "값": _as_number(result.tree_equivalent_count)},
            ]
        )

        meta = pd.DataFrame(
            [
                {"항목": "엔진 버전", "값": result.engine_version},
                {"항목": "계수 세트 버전", "값": result.coefficient_set_version},
                {"항목": "생성 시각 (UTC)", "값": generated_at},
            ]
        )

        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            inputs.to_excel(writer, index=False, sheet_name="입력값")
            outputs.to_excel(writer, index=False, sheet_name="결과")
            meta.to_excel(writer, index=False, sheet_name="메타데이터")

            # --- 간단 서식(가독성 + 숫자 포맷) ---
            from openpyxl.styles import Alignment, Font

            def style_sheet(sheet_name: str, formats: dict[str, str] | None = None) -> None:
                ws = writer.sheets[sheet_name]
                ws.freeze_panes = "A2"

                # Column widths
                ws.column_dimensions["A"].width = 24
                ws.column_dimensions["B"].width = 28

                # Header style
                header_font = Font(bold=True)
                for cell in ws[1]:
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")

                # Body alignment
                for r in range(2, ws.max_row + 1):
                    ws.cell(row=r, column=1).alignment = Alignment(horizontal="left")
                    ws.cell(row=r, column=2).alignment = Alignment(horizontal="left")

                # Number formats by item label
                if formats:
                    for r in range(2, ws.max_row + 1):
                        item = ws.cell(row=r, column=1).value
                        if item in formats:
                            ws.cell(row=r, column=2).number_format = formats[item]

            style_sheet(
                "입력값",
                formats={
                    "옥상 면적(㎡)": "#,##0.00",
                    "녹화 비율": "0.00%",
                    "기준 표면 온도(℃)": "0.0",
                },
            )
            style_sheet(
                "결과",
                formats={
                    "녹화 면적(㎡)": "#,##0.00",
                    "CO₂ 흡수(kg/년)": "#,##0.00",
                    "온도 저감(℃)": "0.00",
                    "녹화 후 표면 온도(℃)": "0.0",
                    "소나무 환산(그루)": "#,##0",
                },
            )
            style_sheet("메타데이터")

        return buf.getvalue(), "okssangimong_result.xlsx"
